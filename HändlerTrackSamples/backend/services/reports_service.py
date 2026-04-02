"""
Servicio de Reportes
Genera reportes en diferentes formatos (Excel, JSON, PDF)
"""

import io
import os
from typing import Optional, Dict, Any, List
from datetime import datetime, date
from decimal import Decimal


class ReportsService:
    """
    Servicio para generar reportes del sistema.

    Tipos de reportes:
    - Inventario: Estado actual del inventario
    - Movimientos: Historial de entradas y salidas
    - Ocupación: Estado de ocupación de anaqueles
    - Vencimientos: Reporte de muestras próximas a vencer
    """

    @staticmethod
    def generar_reporte_inventario(db: Any) -> Dict[str, Any]:
        """
        Genera reporte de inventario actual.

        Args:
            db: Sesión de base de datos

        Returns:
            Diccionario con datos del reporte
        """
        from models.sample import Sample
        from models.proveedor import Proveedor

        # Estadísticas generales
        total_muestras = db.query(Sample).count()
        muestras_activas = db.query(Sample).filter(Sample.estado == "activa").count()
        muestras_agotadas = db.query(Sample).filter(Sample.estado == "agotada").count()
        muestras_vencidas = db.query(Sample).filter(Sample.estado == "vencida").count()

        # Cantidad total en gramos
        total_gramos = (
            db.query(Sample)
            .filter(Sample.estado == "activa")
            .with_entities(db.func.sum(Sample.cantidad_gramos))
            .scalar()
            or 0
        )

        # Por línea de negocio
        from sqlalchemy import func

        por_linea = (
            db.query(
                Sample.linea_negocio,
                func.count(Sample.id).label("cantidad"),
                func.sum(Sample.cantidad_gramos).label("gramos"),
            )
            .filter(Sample.estado == "activa")
            .group_by(Sample.linea_negocio)
            .all()
        )

        # Por proveedor
        por_proveedor = (
            db.query(
                Proveedor.nombre,
                func.count(Sample.id).label("cantidad"),
                func.sum(Sample.cantidad_gramos).label("gramos"),
            )
            .join(Sample, Sample.proveedor_id == Proveedor.id)
            .filter(Sample.estado == "activa")
            .group_by(Proveedor.nombre)
            .all()
        )

        return {
            "success": True,
            "reporte": "inventario",
            "fecha_generacion": datetime.now().isoformat(),
            "resumen": {
                "total_muestras": total_muestras,
                "muestras_activas": muestras_activas,
                "muestras_agotadas": muestras_agotadas,
                "muestras_vencidas": muestras_vencidas,
                "total_gramos": float(total_gramos),
            },
            "por_linea": [
                {"linea": r[0], "cantidad": r[1], "gramos": float(r[2] or 0)}
                for r in por_linea
            ],
            "por_proveedor": [
                {"proveedor": r[0], "cantidad": r[1], "gramos": float(r[2] or 0)}
                for r in por_proveedor
            ],
        }

    @staticmethod
    def generar_reporte_movimientos(
        db: Any, fecha_inicio: Optional[date] = None, fecha_fin: Optional[date] = None
    ) -> Dict[str, Any]:
        """
        Genera reporte de movimientos.

        Args:
            db: Sesión de base de datos
            fecha_inicio: Fecha inicial del período
            fecha_fin: Fecha final del período

        Returns:
            Diccionario con datos del reporte
        """
        from models.movement import Movement

        query = db.query(Movement)

        if fecha_inicio:
            query = query.filter(Movement.fecha_movimiento >= fecha_inicio)
        if fecha_fin:
            query = query.filter(Movement.fecha_movimiento <= fecha_fin)

        movimientos = query.order_by(Movement.fecha_movimiento.desc()).all()

        # Estadísticas
        entradas = [m for m in movimientos if m.tipo == "ENTRADA"]
        salidas = [m for m in movimientos if m.tipo == "SALIDA"]
        reubicaciones = [m for m in movimientos if m.tipo == "REUBICACION"]

        return {
            "success": True,
            "reporte": "movimientos",
            "fecha_generacion": datetime.now().isoformat(),
            "periodo": {
                "inicio": str(fecha_inicio) if fecha_inicio else "desde inicio",
                "fin": str(fecha_fin) if fecha_fin else "hasta hoy",
            },
            "resumen": {
                "total_movimientos": len(movimientos),
                "entradas": len(entradas),
                "salidas": len(salidas),
                "reubicaciones": len(reubicaciones),
            },
            "movimientos": [
                {
                    "id": m.id,
                    "tipo": m.tipo,
                    "muestra_id": m.sample_id,
                    "cantidad_gramos": float(m.cantidad_gramos)
                    if m.cantidad_gramos
                    else 0,
                    "fecha": str(m.fecha_movimiento),
                    "observaciones": m.observaciones,
                }
                for m in movimientos[:100]  # Limitar a 100 registros
            ],
        }

    @staticmethod
    def generar_reporte_ocupacion(db: Any) -> Dict[str, Any]:
        """
        Genera reporte de ocupación de anaqueles.

        Args:
            db: Sesión de base de datos

        Returns:
            Diccionario con datos del reporte
        """
        from models.anaquel import Anaquel
        from models.hilera import Hilera
        from models.linea import Linea
        from sqlalchemy import func

        anaqueles = db.query(Anaquel).all()

        reporte_anaqueles = []

        for anaquel in anaqueles:
            # Contar hileras
            total_hileras = (
                db.query(Hilera).filter(Hilera.anaquel_id == anaquel.id).count()
            )
            hileras_ocupadas = (
                db.query(Hilera)
                .filter(Hilera.anaquel_id == anaquel.id, Hilera.estado == "ocupado")
                .count()
            )

            # Calcular capacidad
            capacidad_gramos = (
                db.query(func.sum(Hilera.capacidad_max))
                .filter(Hilera.anaquel_id == anaquel.id)
                .scalar()
                or 0
            )

            linea_nombre = anaquel.linea.nombre if anaquel.linea else "N/A"

            reporte_anaqueles.append(
                {
                    "id": anaquel.id,
                    "nombre": anaquel.nombre,
                    "linea": linea_nombre,
                    "total_hileras": total_hileras,
                    "hileras_ocupadas": hileras_ocupadas,
                    "hileras_disponibles": total_hileras - hileras_ocupadas,
                    "porcentaje_ocupacion": round(
                        hileras_ocupadas / total_hileras * 100, 1
                    )
                    if total_hileras > 0
                    else 0,
                    "niveles": anaquel.niveles,
                    "hileras_por_nivel": anaquel.hileras_por_nivel,
                }
            )

        # Resumen por línea
        por_linea = {}
        for anaquel_data in reporte_anaqueles:
            linea = anaquel_data["linea"]
            if linea not in por_linea:
                por_linea[linea] = {"total": 0, "ocupadas": 0, "disponibles": 0}
            por_linea[linea]["total"] += anaquel_data["total_hileras"]
            por_linea[linea]["ocupadas"] += anaquel_data["hileras_ocupadas"]
            por_linea[linea]["disponibles"] += anaquel_data["hileras_disponibles"]

        return {
            "success": True,
            "reporte": "ocupacion",
            "fecha_generacion": datetime.now().isoformat(),
            "total_anaqueles": len(reporte_anaqueles),
            "resumen_por_linea": [
                {
                    "linea": k,
                    "total": v["total"],
                    "ocupadas": v["ocupadas"],
                    "disponibles": v["disponibles"],
                }
                for k, v in por_linea.items()
            ],
            "anaqueles": reporte_anaqueles,
        }

    @staticmethod
    def generar_reporte_vencimientos(
        db: Any, dias_adelante: int = 90
    ) -> Dict[str, Any]:
        """
        Genera reporte de muestras próximas a vencer.

        Args:
            db: Sesión de base de datos
            dias_adelante: Días hacia adelante a considerar

        Returns:
            Diccionario con datos del reporte
        """
        from models.sample import Sample
        from datetime import timedelta

        hoy = date.today()
        hasta = hoy + timedelta(days=dias_adelante)

        # Muestras próximas a vencer
        proximas = (
            db.query(Sample)
            .filter(
                Sample.estado == "activa",
                Sample.fecha_vencimiento.isnot(None),
                Sample.fecha_vencimiento >= hoy,
                Sample.fecha_vencimiento <= hasta,
            )
            .order_by(Sample.fecha_vencimiento.asc())
            .all()
        )

        # Muestras ya vencidas
        vencidas = (
            db.query(Sample)
            .filter(
                Sample.estado == "activa",
                Sample.fecha_vencimiento.isnot(None),
                Sample.fecha_vencimiento < hoy,
            )
            .order_by(Sample.fecha_vencimiento.asc())
            .all()
        )

        # Agrupar por urgencia
        urgency_groups = {"critico": [], "alto": [], "medio": [], "bajo": []}

        for muestra in proximas:
            dias = (muestra.fecha_vencimiento - hoy).days

            if dias <= 30:
                grupo = "critico"
            elif dias <= 60:
                grupo = "alto"
            elif dias <= 90:
                grupo = "medio"
            else:
                grupo = "bajo"

            urgencia_groups[grupo].append(
                {
                    "id": muestra.id,
                    "nombre": muestra.nombre,
                    "lote": muestra.lote,
                    "cantidad_gramos": float(muestra.cantidad_gramos),
                    "proveedor": muestra.proveedor.nombre
                    if muestra.proveedor
                    else "N/A",
                    "fecha_vencimiento": str(muestra.fecha_vencimiento),
                    "dias_restantes": dias,
                }
            )

        return {
            "success": True,
            "reporte": "vencimientos",
            "fecha_generacion": datetime.now().isoformat(),
            "resumen": {
                "proximas_vencer": len(proximas),
                "ya_vencidas": len(vencidas),
                "critico": len(urgencia_groups["critico"]),
                "alto": len(urgencia_groups["alto"]),
                "medio": len(urgencia_groups["medio"]),
                "bajo": len(urgencia_groups["bajo"]),
            },
            "urgencia": {
                "critico": urgencia_groups["critico"][:20],
                "alto": urgencia_groups["alto"][:20],
                "medio": urgencia_groups["medio"][:20],
                "bajo": urgencia_groups["bajo"][:20],
            },
            "vencidas": [
                {
                    "id": m.id,
                    "nombre": m.nombre,
                    "lote": m.lote,
                    "cantidad_gramos": float(m.cantidad_gramos),
                    "fecha_vencimiento": str(m.fecha_vencimiento),
                    "dias_vencida": (hoy - m.fecha_vencimiento).days,
                }
                for m in vencidas[:20]
            ],
        }

    @staticmethod
    def exportar_excel(reporte_data: Dict[str, Any]) -> bytes:
        """
        Exporta un reporte a formato Excel.

        Args:
            reporte_data: Datos del reporte

        Returns:
            Bytes del archivo Excel
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            return b"Error: openpyxl no instalado"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte"

        # Título
        ws["A1"] = f"Reporte: {reporte_data.get('reporte', 'General')}"
        ws["A1"].font = Font(bold=True, size=14)

        ws["A2"] = f"Fecha: {reporte_data.get('fecha_generacion', '')}"
        ws["A2"].font = Font(size=10)

        # Agregar datos del reporte
        row = 4
        for key, value in reporte_data.items():
            if key not in ["success", "reporte", "fecha_generacion"]:
                ws[f"A{row}"] = str(key)
                ws[f"B{row}"] = str(value)
                row += 1

        # Guardar a bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def generar_reporte_organizacion(db: Any, linea_id: int = None) -> Dict[str, Any]:
        """
        Genera reporte del estado de organización del almacén.

        Args:
            db: Sesión de base de datos
            linea_id: Filtrar por línea específica (opcional)

        Returns:
            Diccionario con datos del reporte de organización
        """
        from services.motor_organizacion import MotorOrganizacion
        from models.linea import Linea

        # Obtener análisis completo
        analisis = MotorOrganizacion.analizar_organizacion_actual(db, linea_id)

        # Obtener resumen por líneas
        resumen_lineas = MotorOrganizacion.obtener_resumen_lineas(db)

        # Obtener hileras con contenido para análisis detallado
        hileras_contenido = MotorOrganizacion.obtener_hileras_con_contenido(
            db, linea_id
        )

        # Clasificar hileras por estado
        hileras_ok = []
        hileras_incompatibles = []
        hileras_vacias = []

        for hilera in hileras_contenido:
            if hilera.get("muestra") is None:
                hileras_vacias.append(hilera)
            elif hilera.get("tiene_incompatibilidad"):
                hileras_incompatibles.append(hilera)
            else:
                hileras_ok.append(hilera)

        # Porcentajes
        total_hileras = len(hileras_contenido)
        pct_ocupadas = (
            ((len(hileras_ok) + len(hileras_incompatibles)) / total_hileras * 100)
            if total_hileras > 0
            else 0
        )
        pct_incompatibles = (
            (len(hileras_incompatibles) / total_hileras * 100)
            if total_hileras > 0
            else 0
        )

        return {
            "success": True,
            "reporte": "organizacion",
            "fecha_generacion": datetime.now().isoformat(),
            "resumen_ejecutivo": {
                "score_organizacion": analisis.get("score_organizacion", 0),
                "total_muestras": analisis.get("total_muestras", 0),
                "muestras_ok": analisis.get("ok_count", 0),
                "muestras_incompatibles": analisis.get("incompatible_count", 0),
                "muestras_sin_clase": analisis.get("sin_clase_peligro", 0),
                "porcentaje_ocupacion": round(pct_ocupadas, 1),
                "porcentaje_incompatibles": round(pct_incompatibles, 1),
            },
            "resumen_por_linea": resumen_lineas.get("lineas", []),
            "incompatibilidades": [
                {
                    "muestra": p.get("nombre"),
                    "clase_peligro": p.get("clase_peligro"),
                    "incompatibilidades": p.get("incompatibilidades", []),
                }
                for p in analisis.get("problemas", [])[:20]
            ],
            "sugerencias": analisis.get("sugerencias", []),
            "detalle_hileras": {
                "total": total_hileras,
                "ok": len(hileras_ok),
                "incompatibles": len(hileras_incompatibles),
                "vacias": len(hileras_vacias),
            },
        }
